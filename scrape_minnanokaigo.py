import requests
from bs4 import BeautifulSoup
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re
import os

# ===== 設定 =====
URLS = [
    "https://www.minnanokaigo.com/facility/000-7662038528/",
    # 複数URLを追加する場合はここに追記
]

OUTPUT_PATH = os.path.join(os.path.expanduser("~"), "Desktop", f"みんなの介護_施設情報_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "ja,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xhtml;q=0.9,*/*;q=0.8",
}


def scrape_facility(url: str) -> dict:
    """みんなの介護の施設ページから情報を取得する"""
    data = {"URL": url}

    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.raise_for_status()
        response.encoding = "utf-8"
        soup = BeautifulSoup(response.text, "html.parser")

        # ---- 施設名 ----
        h1 = soup.find("h1")
        data["施設名"] = h1.get_text(strip=True) if h1 else ""

        # ---- 施設種別 ----
        facility_type_el = soup.select_one(".p-facility-header__type, .c-facility-label, [class*='facility-type']")
        data["施設種別"] = facility_type_el.get_text(strip=True) if facility_type_el else ""

        # ---- 住所 ----
        address_el = soup.select_one("address, [itemprop='address'], .p-facility-info__address, [class*='address']")
        data["住所"] = address_el.get_text(strip=True) if address_el else ""

        # ---- 電話番号 ----
        tel_el = soup.select_one("[itemprop='telephone'], .c-tel, [class*='tel']")
        if not tel_el:
            tel_match = re.search(r'[\d０-９]{2,4}[-－][\d０-９]{2,4}[-－][\d０-９]{4}', soup.get_text())
            data["電話番号"] = tel_match.group() if tel_match else ""
        else:
            data["電話番号"] = tel_el.get_text(strip=True)

        # ---- 定員 ----
        capacity_el = soup.find(string=re.compile(r'定員'))
        if capacity_el:
            parent = capacity_el.find_parent()
            if parent:
                next_el = parent.find_next_sibling()
                data["定員"] = next_el.get_text(strip=True) if next_el else parent.get_text(strip=True)
        else:
            data["定員"] = ""

        # ---- 料金情報（月額・初期費用）----
        fee_section = soup.select_one("[class*='fee'], [class*='price'], [class*='cost'], [class*='ryoukin']")
        if fee_section:
            data["料金情報"] = fee_section.get_text(separator=" ", strip=True)[:300]
        else:
            data["料金情報"] = ""

        # ---- 月額費用 ----
        monthly_match = re.search(r'月額[^円]*?(\d[\d,]+)\s*円', soup.get_text())
        data["月額費用"] = monthly_match.group() if monthly_match else ""

        # ---- 初期費用 ----
        initial_match = re.search(r'(入居一時金|初期費用)[^円]*?(\d[\d,]*)\s*円', soup.get_text())
        data["初期費用"] = initial_match.group() if initial_match else ""

        # ---- 評価・口コミ数 ----
        rating_el = soup.select_one("[itemprop='ratingValue'], [class*='rating'], [class*='score']")
        data["評価"] = rating_el.get_text(strip=True) if rating_el else ""

        review_match = re.search(r'(\d+)\s*件', soup.get_text())
        data["口コミ数"] = review_match.group() if review_match else ""

        # ---- アクセス ----
        access_el = soup.find(string=re.compile(r'アクセス|交通'))
        if access_el:
            parent = access_el.find_parent()
            if parent:
                data["アクセス"] = parent.get_text(strip=True)[:200]
        else:
            data["アクセス"] = ""

        # ---- 運営会社 ----
        company_el = soup.find(string=re.compile(r'運営|法人|事業者'))
        if company_el:
            parent = company_el.find_parent()
            if parent:
                next_el = parent.find_next_sibling()
                data["運営会社"] = (next_el or parent).get_text(strip=True)[:100]
        else:
            data["運営会社"] = ""

        # ---- テーブル形式の詳細情報を全取得 ----
        tables = soup.find_all("table")
        extra_rows = []
        for table in tables:
            for row in table.find_all("tr"):
                cells = row.find_all(["th", "td"])
                if len(cells) >= 2:
                    key = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if key and val:
                        extra_rows.append(f"{key}: {val}")
        data["テーブル詳細"] = "\n".join(extra_rows[:50])  # 最大50行

        # ---- dl/dt/dd 形式の詳細情報 ----
        dl_items = []
        for dl in soup.find_all("dl"):
            dts = dl.find_all("dt")
            dds = dl.find_all("dd")
            for dt, dd in zip(dts, dds):
                k = dt.get_text(strip=True)
                v = dd.get_text(strip=True)
                if k and v:
                    dl_items.append(f"{k}: {v}")
        data["DL詳細"] = "\n".join(dl_items[:50])

        data["取得日時"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data["ステータス"] = "成功"

    except requests.exceptions.HTTPError as e:
        data["ステータス"] = f"HTTPエラー: {e}"
    except Exception as e:
        data["ステータス"] = f"エラー: {e}"

    return data


def save_to_excel(records: list[dict], path: str):
    """取得データをExcelに保存する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "施設情報"

    # ヘッダー定義（順番通りに並べる）
    columns = [
        "施設名", "施設種別", "住所", "電話番号", "定員",
        "月額費用", "初期費用", "料金情報", "評価", "口コミ数",
        "アクセス", "運営会社", "URL", "取得日時", "ステータス",
        "テーブル詳細", "DL詳細"
    ]

    # スタイル定義
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ヘッダー行
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    # データ行
    for row_idx, record in enumerate(records, start=2):
        row_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(columns, start=1):
            value = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if row_fill.fill_type != "none":
                cell.fill = row_fill

    # 列幅自動調整
    col_widths = {
        "施設名": 30, "施設種別": 20, "住所": 35, "電話番号": 18, "定員": 10,
        "月額費用": 20, "初期費用": 20, "料金情報": 40, "評価": 10, "口コミ数": 12,
        "アクセス": 35, "運営会社": 25, "URL": 50, "取得日時": 20, "ステータス": 15,
        "テーブル詳細": 60, "DL詳細": 60
    }
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    # ウィンドウ枠固定（1行目固定）
    ws.freeze_panes = "A2"

    # オートフィルター
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"\nExcel出力完了: {path}")


# ===== メイン処理 =====
if __name__ == "__main__":
    print(f"=== みんなの介護 施設情報スクレイパー ===")
    print(f"対象URL数: {len(URLS)}\n")

    records = []
    for i, url in enumerate(URLS, start=1):
        print(f"[{i}/{len(URLS)}] 取得中: {url}")
        data = scrape_facility(url)
        records.append(data)
        print(f"  施設名: {data.get('施設名', '(不明)')}")
        print(f"  住所:   {data.get('住所', '(不明)')}")
        print(f"  ステータス: {data.get('ステータス', '')}")

        if i < len(URLS):
            time.sleep(1.5)

    save_to_excel(records, OUTPUT_PATH)
    print(f"\n取得件数: {len(records)}件")
